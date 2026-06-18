"""Generate consolidated reports from all_experiments_summary_*.json files.

Outputs into reports/ with a date stamp:
- tabela_completa_variantes_todos_reruns_<date>.{csv,xlsx,md}
- resumo_resultados_finais_<date>.csv
- relatorio_resultados_finais_<date>.{md,html,pdf}
"""
from __future__ import annotations

import json
import sys
from datetime import date
from pathlib import Path

import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak,
)


WORKSPACE = Path(__file__).resolve().parent.parent
OUTPUTS = WORKSPACE / "outputs"
REPORTS = WORKSPACE / "reports"

SUMMARY_FILES = [
    ("Baseline", "all_experiments_summary.json"),
    ("Rerun 1", "all_experiments_summary_rerun1_imagenet_frozen.json"),
    ("Rerun 2", "all_experiments_summary_rerun2_imagenet_unfrozen.json"),
    ("Rerun 3", "all_experiments_summary_rerun3_controlled_mix_imagenet_unfrozen.json"),
    ("Rerun 4", "all_experiments_summary_rerun4_from_scratch_unfrozen.json"),
    ("Rerun 5", "all_experiments_summary_rerun5_from_scratch_unfrozen_adamw.json"),
    ("Rerun 6 (5:1)", "all_experiments_summary_rerun6_mix_5to1_from_scratch_unfrozen_adamw.json"),
    ("Rerun 7 (2:1)", "all_experiments_summary_rerun7_mix_2to1_from_scratch_unfrozen_adamw.json"),
]


def load_rows() -> list[dict]:
    rows: list[dict] = []
    for run_label, filename in SUMMARY_FILES:
        path = OUTPUTS / filename
        if not path.exists():
            print(f"[WARN] Missing: {path}", file=sys.stderr)
            continue
        payload = json.loads(path.read_text(encoding="utf-8"))
        run_key = path.stem
        for experiment in payload:
            exp_name = experiment["experiment_name"]
            for variant in experiment["variant_results"]:
                rows.append({
                    "RunKey": run_key,
                    "Run": run_label,
                    "Experiment": exp_name,
                    "Variant": variant["variant_name"],
                    "MethodologyLabel": variant.get("methodology_label", ""),
                    "ResultsLabel": variant.get("results_label", ""),
                    "ValidationAccuracy": variant.get("validation_accuracy"),
                    "TestAccuracy": variant.get("test_accuracy"),
                    "Interrupted": variant.get("interrupted", False),
                    "OutputDir": variant.get("output_dir", ""),
                })
    return rows


def fmt_acc(v) -> str:
    if v is None:
        return "-"
    return f"{float(v):.4f}"


def build_summary(df: pd.DataFrame) -> pd.DataFrame:
    grouped = df.groupby(["Run", "Experiment"], sort=False)
    rows = []
    for (run, exp), g in grouped:
        if g["TestAccuracy"].notna().any():
            best_t = g.loc[g["TestAccuracy"].idxmax()]
            best_t_variants = g.loc[g["TestAccuracy"] == best_t["TestAccuracy"], "Variant"].tolist()
            best_t_variant = "/".join(v.replace("Variante_", "") for v in best_t_variants)
            best_t_val = best_t["TestAccuracy"]
        else:
            best_t_variant, best_t_val = "-", None
        if g["ValidationAccuracy"].notna().any():
            best_v = g.loc[g["ValidationAccuracy"].idxmax()]
            best_v_variants = g.loc[g["ValidationAccuracy"] == best_v["ValidationAccuracy"], "Variant"].tolist()
            best_v_variant = "/".join(v.replace("Variante_", "") for v in best_v_variants)
            best_v_val = best_v["ValidationAccuracy"]
        else:
            best_v_variant, best_v_val = "-", None
        mean_test = g["TestAccuracy"].mean() if g["TestAccuracy"].notna().any() else None
        mean_val = g["ValidationAccuracy"].mean() if g["ValidationAccuracy"].notna().any() else None
        rows.append({
            "Run": run,
            "Experiment": exp,
            "BestTestVariant": best_t_variant,
            "BestTestAccuracy": best_t_val,
            "BestValidationVariant": best_v_variant,
            "BestValidationAccuracy": best_v_val,
            "MeanTestAccuracy": mean_test,
            "MeanValidationAccuracy": mean_val,
        })
    return pd.DataFrame(rows)


def build_top10(df: pd.DataFrame) -> pd.DataFrame:
    valid = df[df["TestAccuracy"].notna()].copy()
    valid = valid.sort_values(
        by=["TestAccuracy", "ValidationAccuracy"],
        ascending=[False, False],
    ).head(10).reset_index(drop=True)
    valid.insert(0, "Posicao", range(1, len(valid) + 1))
    return valid[[
        "Posicao", "Run", "Experiment", "Variant",
        "ValidationAccuracy", "TestAccuracy",
    ]]


DROP_COLS = ("MethodologyLabel", "ResultsLabel", "Interrupted", "OutputDir")


def interleave_experiment_means(df: pd.DataFrame) -> pd.DataFrame:
    """Para cada (Run, Experimento): listar todas as variantes e apos
    o ultimo item do grupo, inserir uma linha MEDIA com a media
    daquele experimento naquele run especifico."""
    out: list[dict] = []
    for (run_key, run, exp), g in df.groupby(["RunKey", "Run", "Experiment"], sort=False):
        for _, r in g.iterrows():
            out.append(r.to_dict())
        out.append({
            "RunKey": run_key,
            "Run": run,
            "Experiment": exp,
            "Variant": "MEDIA",
            "ValidationAccuracy": g["ValidationAccuracy"].mean(),
            "TestAccuracy": g["TestAccuracy"].mean(),
        })
    return pd.DataFrame(out)


def export_csv_xlsx(df_full: pd.DataFrame, df_summary: pd.DataFrame, df_top10: pd.DataFrame, stamp: str):
    df_full_slim = df_full.drop(columns=list(DROP_COLS))
    df_with_means = interleave_experiment_means(df_full_slim)

    csv_path = REPORTS / f"tabela_completa_variantes_todos_reruns_{stamp}.csv"
    df_csv = df_with_means.copy()
    for col in ("ValidationAccuracy", "TestAccuracy"):
        df_csv[col] = df_csv[col].apply(lambda v: "" if pd.isna(v) else f"{v}")
    df_csv.to_csv(csv_path, index=False, encoding="utf-8-sig")

    xlsx_path = REPORTS / f"tabela_completa_variantes_todos_reruns_{stamp}.xlsx"
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df_with_means.to_excel(writer, sheet_name="Completo", index=False)
        workbook = writer.book
        ws = workbook["Completo"]
        # ajustar largura das colunas
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max((len(str(c.value or "")) for c in column_cells), default=10)
            ws.column_dimensions[letter].width = min(max_len + 2, 60)
        # negrito + fundo cinza nas linhas MEDIA
        from openpyxl.styles import Font, PatternFill
        media_font = Font(bold=True)
        media_fill = PatternFill("solid", fgColor="E2E8F0")
        variant_col = list(df_with_means.columns).index("Variant") + 1
        for row_index in range(2, ws.max_row + 1):
            if ws.cell(row=row_index, column=variant_col).value == "MEDIA":
                for col_index in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_index, column=col_index)
                    cell.font = media_font
                    cell.fill = media_fill

    resumo_csv = REPORTS / f"resumo_resultados_finais_{stamp}.csv"
    df_summary.to_csv(resumo_csv, index=False, encoding="utf-8-sig")
    return csv_path, xlsx_path, resumo_csv


def build_markdown(df_full: pd.DataFrame, df_summary: pd.DataFrame, df_top10: pd.DataFrame, stamp: str) -> str:
    lines = []
    lines.append("# Relatorio Final de Resultados")
    lines.append("")
    lines.append(f"Data de consolidacao: {stamp}")
    lines.append("")
    lines.append("## Fontes usadas")
    lines.append("")
    for run, fn in SUMMARY_FILES:
        if (OUTPUTS / fn).exists():
            lines.append(f"- `outputs/{fn}` ({run})")
    lines.append("")

    lines.append("## Resumo executivo")
    lines.append("")
    lines.append(
        "- `Rerun 5` (from scratch + AdamW, unfrozen backbone, 300 epocas, patience 50) "
        "concluiu todos os 4 experimentos com 7 variantes cada. Nao usa pesos da ImageNet."
    )
    rerun5 = df_full[df_full["Run"] == "Rerun 5"]
    if not rerun5.empty:
        best_row = rerun5.loc[rerun5["TestAccuracy"].idxmax()]
        lines.append(
            f"- Melhor teste geral do Rerun 5: `{best_row['Experiment']}` / `{best_row['Variant']}` "
            f"= `{best_row['TestAccuracy']:.4f}` (validacao `{best_row['ValidationAccuracy']:.4f}`)."
        )
        worst_exp = rerun5.groupby("Experiment")["TestAccuracy"].mean().idxmin()
        worst_mean = rerun5.groupby("Experiment")["TestAccuracy"].mean().min()
        best_exp = rerun5.groupby("Experiment")["TestAccuracy"].mean().idxmax()
        best_mean = rerun5.groupby("Experiment")["TestAccuracy"].mean().max()
        lines.append(f"- Experimento com melhor media Rerun 5: `{best_exp}` = `{best_mean:.4f}`.")
        lines.append(f"- Experimento com pior media Rerun 5: `{worst_exp}` = `{worst_mean:.4f}`.")
    lines.append("")

    lines.append("## Tabela 1. Melhor resultado por experimento e run")
    lines.append("")
    lines.append("| Run | Experimento | Melhor teste | Melhor validacao | Media teste |")
    lines.append("| --- | --- | --- | --- | ---: |")
    for _, r in df_summary.iterrows():
        bt = f"`Variante_{r['BestTestVariant']}` = `{fmt_acc(r['BestTestAccuracy'])}`" if r["BestTestVariant"] != "-" else "-"
        bv = f"`Variante_{r['BestValidationVariant']}` = `{fmt_acc(r['BestValidationAccuracy'])}`" if r["BestValidationVariant"] != "-" else "-"
        mt = fmt_acc(r["MeanTestAccuracy"])
        lines.append(f"| {r['Run']} | `{r['Experiment']}` | {bt} | {bv} | `{mt}` |")
    lines.append("")

    lines.append("## Tabela 2. Top 10 modelos no teste")
    lines.append("")
    lines.append("| Posicao | Run | Experimento | Variante | Validacao | Teste |")
    lines.append("| ---: | --- | --- | --- | ---: | ---: |")
    for _, r in df_top10.iterrows():
        lines.append(
            f"| {r['Posicao']} | {r['Run']} | `{r['Experiment']}` | `{r['Variant']}` | "
            f"`{fmt_acc(r['ValidationAccuracy'])}` | `{fmt_acc(r['TestAccuracy'])}` |"
        )
    lines.append("")

    lines.append("## Tabela 3. Rerun 5 completo (from scratch + AdamW)")
    lines.append("")
    rerun5 = df_full[df_full["Run"] == "Rerun 5"]
    for exp_name in rerun5["Experiment"].unique():
        sub = rerun5[rerun5["Experiment"] == exp_name]
        lines.append(f"### `{exp_name}`")
        lines.append("")
        lines.append("| Variante | Validacao | Teste |")
        lines.append("| --- | ---: | ---: |")
        for _, r in sub.iterrows():
            lines.append(
                f"| `{r['Variant']}` | `{fmt_acc(r['ValidationAccuracy'])}` | `{fmt_acc(r['TestAccuracy'])}` |"
            )
        lines.append("")

    return "\n".join(lines)


def build_html(md_text: str, stamp: str) -> str:
    try:
        import markdown as md  # type: ignore
        body = md.markdown(md_text, extensions=["tables", "fenced_code"])
    except Exception:
        body = _simple_md_to_html(md_text)
    html = f"""<!DOCTYPE html>
<html lang="pt-br">
<head>
<meta charset="utf-8">
<title>Relatorio Final de Resultados - {stamp}</title>
<style>
body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; margin: 28px; color: #222; max-width: 1100px; }}
h1 {{ color: #1a365d; border-bottom: 2px solid #1a365d; padding-bottom: 6px; }}
h2 {{ color: #2d3748; border-bottom: 1px solid #cbd5e0; padding-bottom: 4px; margin-top: 28px; }}
h3 {{ color: #2d3748; }}
table {{ border-collapse: collapse; width: 100%; margin: 12px 0; font-size: 14px; }}
th, td {{ border: 1px solid #cbd5e0; padding: 6px 10px; text-align: left; }}
th {{ background: #edf2f7; }}
code {{ background: #f1f5f9; padding: 1px 5px; border-radius: 4px; font-size: 13px; }}
tr:nth-child(even) {{ background: #fafafa; }}
</style>
</head>
<body>
{body}
</body>
</html>
"""
    return html


def _simple_md_to_html(md_text: str) -> str:
    # Fallback minimal converter supporting headers, tables, paragraphs, code.
    out = []
    lines = md_text.split("\n")
    i = 0
    while i < len(lines):
        line = lines[i]
        stripped = line.strip()
        if stripped.startswith("### "):
            out.append(f"<h3>{_inline(stripped[4:])}</h3>")
            i += 1
        elif stripped.startswith("## "):
            out.append(f"<h2>{_inline(stripped[3:])}</h2>")
            i += 1
        elif stripped.startswith("# "):
            out.append(f"<h1>{_inline(stripped[2:])}</h1>")
            i += 1
        elif stripped.startswith("| "):
            table_lines = []
            while i < len(lines) and lines[i].strip().startswith("|"):
                table_lines.append(lines[i])
                i += 1
            out.append(_table_to_html(table_lines))
        elif stripped.startswith("- "):
            items = []
            while i < len(lines) and lines[i].strip().startswith("- "):
                items.append(f"<li>{_inline(lines[i].strip()[2:])}</li>")
                i += 1
            out.append("<ul>" + "".join(items) + "</ul>")
        elif stripped == "":
            i += 1
        else:
            out.append(f"<p>{_inline(stripped)}</p>")
            i += 1
    return "\n".join(out)


def _inline(text: str) -> str:
    import html as h
    text = h.escape(text)
    # simple backtick -> code
    parts = []
    in_code = False
    for tok in text.split("`"):
        if in_code:
            parts.append(f"<code>{tok}</code>")
        else:
            parts.append(tok)
        in_code = not in_code
    return "".join(parts)


def _table_to_html(lines: list[str]) -> str:
    rows = []
    for ln in lines:
        cells = [c.strip() for c in ln.strip().strip("|").split("|")]
        rows.append(cells)
    if len(rows) < 2:
        return ""
    header = rows[0]
    align = rows[1]
    body = rows[2:]
    out = ["<table><thead><tr>"]
    for h_ in header:
        out.append(f"<th>{_inline(h_)}</th>")
    out.append("</tr></thead><tbody>")
    for row in body:
        out.append("<tr>")
        for cell in row:
            out.append(f"<td>{_inline(cell)}</td>")
        out.append("</tr>")
    out.append("</tbody></table>")
    return "".join(out)


def export_pdf(df_full: pd.DataFrame, df_summary: pd.DataFrame, df_top10: pd.DataFrame, stamp: str) -> Path:
    pdf_path = REPORTS / f"relatorio_resultados_finais_{stamp}.pdf"
    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=landscape(A4),
        leftMargin=14 * mm, rightMargin=14 * mm,
        topMargin=14 * mm, bottomMargin=14 * mm,
    )
    styles = getSampleStyleSheet()
    h1 = styles["Heading1"]
    h2 = styles["Heading2"]
    h3 = styles["Heading3"]
    body = styles["BodyText"]
    mono = ParagraphStyle("mono", parent=body, fontName="Courier", fontSize=8)

    flow = []
    flow.append(Paragraph("Relatorio Final de Resultados", h1))
    flow.append(Paragraph(f"Data de consolidacao: {stamp}", body))
    flow.append(Spacer(1, 6))

    flow.append(Paragraph("Tabela 1. Melhor resultado por experimento e run", h2))
    data1 = [["Run", "Experimento", "Melhor teste", "Melhor validacao", "Media teste"]]
    for _, r in df_summary.iterrows():
        bt = f"Var_{r['BestTestVariant']} = {fmt_acc(r['BestTestAccuracy'])}" if r["BestTestVariant"] != "-" else "-"
        bv = f"Var_{r['BestValidationVariant']} = {fmt_acc(r['BestValidationAccuracy'])}" if r["BestValidationVariant"] != "-" else "-"
        data1.append([r["Run"], r["Experiment"], bt, bv, fmt_acc(r["MeanTestAccuracy"])])
    t1 = Table(data1, repeatRows=1, colWidths=[55, 130, 95, 95, 55])
    t1.setStyle(_table_style())
    flow.append(t1)
    flow.append(Spacer(1, 10))

    flow.append(Paragraph("Tabela 2. Top 10 modelos no teste", h2))
    data2 = [["#", "Run", "Experimento", "Variante", "Validacao", "Teste"]]
    for _, r in df_top10.iterrows():
        data2.append([
            str(r["Posicao"]), r["Run"], r["Experiment"], r["Variant"],
            fmt_acc(r["ValidationAccuracy"]), fmt_acc(r["TestAccuracy"]),
        ])
    t2 = Table(data2, repeatRows=1, colWidths=[20, 55, 140, 75, 55, 55])
    t2.setStyle(_table_style())
    flow.append(t2)
    flow.append(PageBreak())

    flow.append(Paragraph("Tabela 3. Rerun 5 completo (from scratch + AdamW)", h2))
    rerun5 = df_full[df_full["Run"] == "Rerun 5"]
    for exp_name in rerun5["Experiment"].unique():
        flow.append(Paragraph(exp_name, h3))
        sub = rerun5[rerun5["Experiment"] == exp_name]
        data = [["Variante", "Validacao", "Teste"]]
        for _, r in sub.iterrows():
            data.append([r["Variant"], fmt_acc(r["ValidationAccuracy"]), fmt_acc(r["TestAccuracy"])])
        t = Table(data, repeatRows=1, colWidths=[80, 60, 60])
        t.setStyle(_table_style())
        flow.append(t)
        flow.append(Spacer(1, 6))

    flow.append(PageBreak())
    flow.append(Paragraph("Tabela 4. Todas as variantes x runs (completa)", h2))
    data_full = [["Run", "Experimento", "Variante", "Validacao", "Teste"]]
    for _, r in df_full.iterrows():
        data_full.append([
            r["Run"], r["Experiment"], r["Variant"],
            fmt_acc(r["ValidationAccuracy"]), fmt_acc(r["TestAccuracy"]),
        ])
    t_full = Table(data_full, repeatRows=1, colWidths=[55, 140, 75, 55, 55])
    t_full.setStyle(_table_style())
    flow.append(t_full)

    doc.build(flow)
    return pdf_path


def _table_style() -> TableStyle:
    return TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#edf2f7")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1a365d")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fafafa")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ])


def main() -> None:
    REPORTS.mkdir(parents=True, exist_ok=True)
    stamp = date.today().isoformat()
    rows = load_rows()
    if not rows:
        raise SystemExit("Nenhum summary encontrado.")
    df_full = pd.DataFrame(rows)
    df_summary = build_summary(df_full)
    df_top10 = build_top10(df_full)

    csv_path, xlsx_path, resumo_csv = export_csv_xlsx(df_full, df_summary, df_top10, stamp)

    md_text = build_markdown(df_full, df_summary, df_top10, stamp)
    md_path = REPORTS / f"relatorio_resultados_finais_{stamp}.md"
    md_path.write_text(md_text, encoding="utf-8")

    html_text = build_html(md_text, stamp)
    html_path = REPORTS / f"relatorio_resultados_finais_{stamp}.html"
    html_path.write_text(html_text, encoding="utf-8")

    md_table_path = REPORTS / f"tabela_completa_variantes_todos_reruns_{stamp}.md"
    md_table_lines = [
        "# Tabela completa de variantes por run",
        "",
        f"Data: {stamp}",
        "",
        "| Run | Experimento | Variante | Validacao | Teste | Interrupted |",
        "| --- | --- | --- | ---: | ---: | :---: |",
    ]
    for _, r in df_full.iterrows():
        md_table_lines.append(
            f"| {r['Run']} | `{r['Experiment']}` | `{r['Variant']}` | "
            f"`{fmt_acc(r['ValidationAccuracy'])}` | `{fmt_acc(r['TestAccuracy'])}` | "
            f"{'sim' if r['Interrupted'] else 'nao'} |"
        )
    md_table_path.write_text("\n".join(md_table_lines), encoding="utf-8")

    pdf_path = export_pdf(df_full, df_summary, df_top10, stamp)

    print("Gerados:")
    for p in (csv_path, xlsx_path, resumo_csv, md_path, html_path, md_table_path, pdf_path):
        print(f"  {p}")


if __name__ == "__main__":
    main()
